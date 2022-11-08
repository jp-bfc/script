import openpyxl
from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Color, PatternFill
import pyodbc
import pandas as pnd
import configparser

config = configparser.ConfigParser()
config.read('script_configs.ini')
username = config['DEFAULT']['user']
password = config['DEFAULT']['password']
server = config['DEFAULT']['server']
port = config['DEFAULT']['port']
database = config['DEFAULT']['database']
cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';PORT='+port+';DATABASE='+database+';UID='+username+';PWD='+ password)
cursor = cnxn.cursor()



# query ="""
# select obj.F1024 as Dept, sum(F64) as Qty, sum(F65) as Revenue, sum(F1301) as Cost,
# cast(month(rpt.F254) as int) as 'Month',
# cast(year(rpt.F254) as int) as 'Year',
# sum(F65) - sum(F1301) as Profit,
# 'Margin' =
# CASE
# 	WHEN sum(F65) > 0 THEN ROUND((sum(F65) - sum(F1301))/sum(F65), 3)
# 	WHEN sum(F65) = 0 THEN 0
# END,
# LAG(sum(F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024) as LY_Volume,
# ((sum(F64) - (LAG(sum(rpt.F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024))) as VOL_GRO,
# LAG(sum(rpt.F65), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024) as LY_REV,
# ((sum(F65) - (LAG(sum(rpt.F65), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F65), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024))) as REV_GRO,
# LAG(sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024) as LY_COST,
# ((sum(F1301) - (LAG(sum(rpt.F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024))) as COST_GRO,
# LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024) as LY_PROFIT,
# (((sum(F65) - sum(F1301))-(LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024))) / LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)) as PROFIT_GROWTH,
# LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024) as LY_MARGIN,
# (((ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2))-(LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024))) / LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)) as MARGIN_GROWTH,
# (sum(F65) - sum(F1301)) - LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024) as GrossProfitVariance,
# ROUND((sum(F65)*ROUND((sum(F65) - sum(F1301))/sum(F65), 4)) - (sum(F65)*LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 4), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)), 2) as DueToMargin,
# ROUND((sum(F64) - LAG(sum(F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), obj.F1024)) * (sum(F1301)/sum(F64)), 2) as VolumeVariance
# from STORESQL.dbo.RPT_ITM_M rpt
# inner join
# (select objt.F01, F155, F29, rpc.F1024, sdp.F1022, rpc.F18
# from STORESQL.dbo.OBJ_TAB objt
# inner join (select F18, F1024 from STORESQL.dbo.RPC_TAB) rpc on objt.F18 = rpc.F18
# inner join (select F01, F04 from STORESQL.dbo.POS_TAB) pos on objt.F01 = pos.F01
# inner join (select * from STORESQL.dbo.SDP_TAB) sdp on pos.F04 = sdp.F04
# where (rpc.F18 <> 97 and rpc.F18 <> 23 and rpc.F18 <> 98 and rpc.F18 <> 14)
# ) obj
# on rpt.F01 = obj.F01
# where rpt.F1034 = 3
# group by obj.F1024, rpt.F254
# order by datepart(year, rpt.F254), datepart(month, rpt.F254), F1024
# """




query="""select obj.F1024 as Dept, sum(F64) as Qty, sum(F65) as Revenue, sum(F1301) as Cost,
cast(month(rpt.F254) as int) as 'Month',
cast(day(rpt.F254) as int) as 'Week Ending',
cast(year(rpt.F254) as int) as 'Year',
sum(F65) - sum(F1301) as Profit,
'Margin' =
CASE
	WHEN sum(F65) > 0 THEN ROUND((sum(F65) - sum(F1301))/sum(F65), 4)
	WHEN sum(F65) = 0 THEN 0
END,
AVG(sum(F65)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 3 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Revenue,
AVG(sum(F64)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 3 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Volume,
AVG(ROUND((sum(F65) - sum(F1301))/sum(F65), 4)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 3 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Margin,
AVG(sum(F1301)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 3 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Cost,
AVG(sum(F65) - sum(F1301)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 3 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Profit,
LAG(sum(F64), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_Volume,
((sum(F64) - (LAG(sum(rpt.F64), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F64), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as VOL_GROWTH_MONTH,
LAG(sum(rpt.F65), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_REV,
((sum(F65) - (LAG(sum(rpt.F65), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F65), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as REV_GROWTH_MONTH,
LAG(sum(F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_COST,
((sum(F1301) - (LAG(sum(rpt.F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as COST_GROWTH_MONTH,
LAG(sum(F65) - sum(F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_PROFIT,
(((sum(F65) - sum(F1301))-(LAG(sum(F65) - sum(F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(sum(F65) - sum(F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as PROFIT_GROWTH_MONTH,
LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_MARGIN,
(((ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2))-(LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65))*100, 2), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as MARGIN_GROWTH_MONTH,
(sum(F65) - sum(F1301)) - LAG(sum(F65) - sum(F1301), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as GrossProfitVarianceMONTH,
ROUND((sum(F65)*ROUND((sum(F65) - sum(F1301))/sum(F65), 1)) - (sum(F65)*LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 1), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)), 2) as DueToMarginMONTH,
ROUND((sum(F64) - LAG(sum(F64), 1) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) * (sum(F1301)/sum(F64)), 2) as VolumeVarianceMONTH,
LAG(sum(F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_Volume,
((sum(F64) - (LAG(sum(rpt.F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as VOL_GRO,
LAG(sum(rpt.F65), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_REV,
((sum(F65) - (LAG(sum(rpt.F65), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F65), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as REV_GRO,
LAG(sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_COST,
((sum(F1301) - (LAG(sum(rpt.F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as COST_GRO,
LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_PROFIT,
(((sum(F65) - sum(F1301))-(LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as PROFIT_GROWTH,
LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_MARGIN,
(((ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2))-(LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65))*100, 2), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as MARGIN_GROWTH,
(sum(F65) - sum(F1301)) - LAG(sum(F65) - sum(F1301), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as GrossProfitVariance,
ROUND((sum(F65)*ROUND((sum(F65) - sum(F1301))/sum(F65), 1)) - (sum(F65)*LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 1), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)), 2) as DueToMargin,
ROUND((sum(F64) - LAG(sum(F64), 12) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) * (sum(F1301)/sum(F64)), 2) as VolumeVariance
from STORESQL.dbo.RPT_ITM_M rpt
inner join
(select objt.F01, F155, F29, rpc.F1024, sdp.F1022, rpc.F18
from STORESQL.dbo.OBJ_TAB objt
inner join (select F18, F1024 from STORESQL.dbo.RPC_TAB) rpc on objt.F18 = rpc.F18
inner join (select F01, F04 from STORESQL.dbo.POS_TAB) pos on objt.F01 = pos.F01
inner join (select * from STORESQL.dbo.SDP_TAB) sdp on pos.F04 = sdp.F04
where (rpc.F18 <> 97 and rpc.F18 <> 23 and rpc.F18 <> 98 and rpc.F18 <> 14)
) obj
on rpt.F01 = obj.F01
where rpt.F1034 = 3
group by obj.F1024, rpt.F254
order by datepart(year, rpt.F254) desc, datepart(month, rpt.F254) desc, datepart(day, rpt.F254) desc, F1024"""




# query="""
# select obj.F1024 as Dept, sum(F64) as Qty, sum(F65) as Revenue, sum(F1301) as Cost,
# cast(month(rpt.F254) as int) as 'Month',
# cast(day(rpt.F254) as int) as 'Week Ending',
# cast(year(rpt.F254) as int) as 'Year',
# sum(F65) - sum(F1301) as Profit,
# 'Margin' =
# CASE
# 	WHEN sum(F65) > 0 THEN ROUND((sum(F65) - sum(F1301))/sum(F65), 4)
# 	WHEN sum(F65) = 0 THEN 0
# END,
# AVG(sum(F65)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 12 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Revenue,
# AVG(sum(F64)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 12 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Volume,
# AVG((sum(F65) - sum(F1301))/sum(F65)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 12 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Margin,
# AVG(sum(F1301)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 12 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Cost,
# AVG(sum(F65) - sum(F1301)) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024 ROWS BETWEEN 12 PRECEDING AND 1 PRECEDING) as Ninety_Trailing_Profit,
# LAG(sum(F64), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_Volume,
# ((sum(F64) - (LAG(sum(rpt.F64), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F64), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as VOL_GROWTH_MONTH,
# LAG(sum(rpt.F65), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_REV,
# ((sum(F65) - (LAG(sum(rpt.F65), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F65), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as REV_GROWTH_MONTH,
# LAG(sum(F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_COST,
# ((sum(F1301) - (LAG(sum(rpt.F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as COST_GROWTH_MONTH,
# LAG(sum(F65) - sum(F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_PROFIT,
# (((sum(F65) - sum(F1301))-(LAG(sum(F65) - sum(F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(sum(F65) - sum(F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as PROFIT_GROWTH_MONTH,
# LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LM_MARGIN,
# (((ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2))-(LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 2), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65))*100, 2), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as MARGIN_GROWTH_MONTH,
# (sum(F65) - sum(F1301)) - LAG(sum(F65) - sum(F1301), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as GrossProfitVarianceMONTH,
# ROUND((sum(F65)*ROUND((sum(F65) - sum(F1301))/sum(F65), 4)) - (sum(F65)*LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 4), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)), 2) as DueToMarginMONTH,
# ROUND((sum(F64) - LAG(sum(F64), 4) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) * (sum(F1301)/sum(F64)), 2) as VolumeVarianceMONTH,
# LAG(sum(F64), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_Volume,
# ((sum(F64) - (LAG(sum(rpt.F64), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F64), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as VOL_GRO,
# LAG(sum(rpt.F65), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_REV,
# ((sum(F65) - (LAG(sum(rpt.F65), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F65), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as REV_GRO,
# LAG(sum(F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_COST,
# ((sum(F1301) - (LAG(sum(rpt.F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)))/ (LAG(sum(rpt.F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as COST_GRO,
# LAG(sum(F65) - sum(F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_PROFIT,
# (((sum(F65) - sum(F1301))-(LAG(sum(F65) - sum(F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) / LAG(sum(F65) - sum(F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) as PROFIT_GROWTH,
# LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 4), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as LY_MARGIN,
# ((ROUND(((sum(F65) - sum(F1301))/sum(F65)), 4))-(LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 4), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024))) as MARGIN_GROWTH,
# (sum(F65) - sum(F1301)) - LAG(sum(F65) - sum(F1301), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024) as GrossProfitVariance,
# ROUND((sum(F65)*ROUND((sum(F65) - sum(F1301))/sum(F65), 4)) - (sum(F65)*LAG(ROUND(((sum(F65) - sum(F1301))/sum(F65)), 4), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)), 2) as DueToMargin,
# ROUND((sum(F64) - LAG(sum(F64), 52) OVER (partition by obj.F1024 ORDER BY datepart(year, rpt.F254) asc, datepart(month, rpt.F254), datepart(day, rpt.F254), obj.F1024)) * (sum(F1301)/sum(F64)), 2) as VolumeVariance
# from STORESQL.dbo.RPT_ITM_W rpt
# inner join
# (select objt.F01, F155, F29, rpc.F1024, sdp.F1022, rpc.F18
# from STORESQL.dbo.OBJ_TAB objt
# inner join (select F18, F1024 from STORESQL.dbo.RPC_TAB) rpc on objt.F18 = rpc.F18
# inner join (select F01, F04 from STORESQL.dbo.POS_TAB) pos on objt.F01 = pos.F01
# inner join (select * from STORESQL.dbo.SDP_TAB) sdp on pos.F04 = sdp.F04
# where (rpc.F18 <> 97 and rpc.F18 <> 23 and rpc.F18 <> 98 and rpc.F18 <> 14)
# ) obj
# on rpt.F01 = obj.F01
# where rpt.F1034 = 3
# group by obj.F1024, rpt.F254
# order by datepart(year, rpt.F254) desc, datepart(month, rpt.F254) desc, datepart(day, rpt.F254) desc, F1024
# """

weeklydata = pnd.read_sql(query, cnxn)

pd = openpyxl.Workbook()

thick_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

greenFill = PatternFill(start_color='0099CC00',
                   end_color='0099CC00',
                   fill_type='solid')

pd.create_sheet("Report")
sheet = pd['Report']
i = 2
sheet.cell(1, 3).value = 'Gross Sales'
sheet.merge_cells('C1:I1')
sheet.cell(1, 3).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 10).value = 'COGS'
sheet.merge_cells('J1:O1')
sheet.cell(1, 10).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 16).value = 'Gross Margin'
sheet.merge_cells('P1:U1')
sheet.cell(1, 16).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 22).value = 'Profit'
sheet.merge_cells('V1:AB1')
sheet.cell(1, 22).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 29).value = 'Volume'
sheet.merge_cells('AC1:AF1')
sheet.cell(1, 29).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 33).value = 'Variance'
sheet.merge_cells('AG1:AI1')
sheet.cell(1, 33).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 36).value = 'Labor $'
sheet.merge_cells('AJ1:AN1')
sheet.cell(1, 36).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 41).value = 'Labor Hours'
sheet.merge_cells('AO1:AS1')
sheet.cell(1, 41).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 46).value = 'Sales $ per Labor Hour'
sheet.merge_cells('AT1:AX1')
sheet.cell(1, 46).alignment = Alignment(horizontal="center", vertical="center")
sheet.cell(1, 51).value = 'Margin-Labor'
sheet.merge_cells('AY1:BC1')
sheet.cell(1, 51).alignment = Alignment(horizontal="center", vertical="center")
# sheet.cell(1, 48).value = 'Trailing AVGs'
# sheet.merge_cells('AV1:AY1')
# sheet.cell(1, 48).alignment = Alignment(horizontal="center", vertical="center")

sheet.cell(1, 3).border = thick_border
sheet.cell(1, 10).border = thick_border
sheet.cell(1, 16).border = thick_border
sheet.cell(1, 22).border = thick_border
sheet.cell(1, 29).border = thick_border
sheet.cell(1, 33).border = thick_border
sheet.cell(1, 36).border = thick_border
sheet.cell(1, 41).border = thick_border
sheet.cell(1, 46).border = thick_border
sheet.cell(1, 51).border = thick_border

#Gross Sales
sheet.cell(2, 3).value = "Actual"
sheet.cell(2, 4).value = "%  Total"
sheet.cell(2, 5).value = "Plan"
sheet.cell(2, 6).value = "% Plan"
sheet.cell(2, 7).value = "90 Day"
sheet.cell(2, 8).value = "LY"
sheet.cell(2, 9).value = "% LY"
#COGS
sheet.cell(2, 10).value = "Actual"
sheet.cell(2, 11).value = "Plan"
sheet.cell(2, 12).value = "% Plan"
sheet.cell(2, 13).value = "90 Day"
sheet.cell(2, 14).value = "LY"
sheet.cell(2, 15).value = "% LY"
#Gross Margin
sheet.cell(2, 16).value = "Actual"
sheet.cell(2, 17).value = "Plan"
sheet.cell(2, 18).value = "% Plan"
sheet.cell(2, 19).value = "90 Day"
sheet.cell(2, 20).value = "LY"
sheet.cell(2, 21).value = "% LY"
#Profit
sheet.cell(2, 22).value = "Actual"
sheet.cell(2, 23).value = "%  Total"
sheet.cell(2, 24).value = "Plan"
sheet.cell(2, 25).value = "% Plan"
sheet.cell(2, 26).value = "90 Day"
sheet.cell(2, 27).value = "LY"
sheet.cell(2, 28).value = "% LY"
#Volume
sheet.cell(2, 29).value = "Volume"
sheet.cell(2, 30).value = "90 Day"
sheet.cell(2, 31).value = "LY"
sheet.cell(2, 32).value = "% LY"
#Variance
sheet.cell(2, 33).value = "Total GP Variance"
sheet.cell(2, 34).value = "Due to Margin %"
sheet.cell(2, 35).value = "Due to Sales Volume"
#Labor $
sheet.cell(2, 36).value = "Actual"
sheet.cell(2, 37).value = "Plan"
sheet.cell(2, 38).value = "% Plan"
sheet.cell(2, 39).value = "LY"
sheet.cell(2, 40).value = "% LY"
#Labor Hours
sheet.cell(2, 41).value = "Actual"
sheet.cell(2, 42).value = "Plan"
sheet.cell(2, 43).value = "% Plan"
sheet.cell(2, 44).value = "LY"
sheet.cell(2, 45).value = "% LY"
#Sales $ per Labor Hour
sheet.cell(2, 46).value = "Actual"
sheet.cell(2, 47).value = "Plan"
sheet.cell(2, 48).value = "% Plan"
sheet.cell(2, 49).value = "LY"
sheet.cell(2, 50).value = "% LY"
#Margin-Labor
sheet.cell(2, 51).value = "Actual"
sheet.cell(2, 52).value = "Plan"
sheet.cell(2, 53).value = "% Plan"
sheet.cell(2, 54).value = "LY"
sheet.cell(2, 55).value = "% LY"

sheet.cell(2, 3).alignment = Alignment(horizontal="center")
sheet.cell(2, 4).alignment = Alignment(horizontal="center")
sheet.cell(2, 5).alignment = Alignment(horizontal="center")
sheet.cell(2, 6).alignment = Alignment(horizontal="center")
sheet.cell(2, 7).alignment = Alignment(horizontal="center")
sheet.cell(2, 8).alignment = Alignment(horizontal="center")
sheet.cell(2, 9).alignment = Alignment(horizontal="center")
sheet.cell(2, 10).alignment = Alignment(horizontal="center")
sheet.cell(2, 11).alignment = Alignment(horizontal="center")
sheet.cell(2, 12).alignment = Alignment(horizontal="center")
sheet.cell(2, 13).alignment = Alignment(horizontal="center")
sheet.cell(2, 14).alignment = Alignment(horizontal="center")
sheet.cell(2, 15).alignment = Alignment(horizontal="center")
sheet.cell(2, 16).alignment = Alignment(horizontal="center")
sheet.cell(2, 17).alignment = Alignment(horizontal="center")
sheet.cell(2, 18).alignment = Alignment(horizontal="center")
sheet.cell(2, 19).alignment = Alignment(horizontal="center")
sheet.cell(2, 20).alignment = Alignment(horizontal="center")
sheet.cell(2, 21).alignment = Alignment(horizontal="center")
sheet.cell(2, 22).alignment = Alignment(horizontal="center")
sheet.cell(2, 23).alignment = Alignment(horizontal="center")
sheet.cell(2, 24).alignment = Alignment(horizontal="center")
sheet.cell(2, 25).alignment = Alignment(horizontal="center")
sheet.cell(2, 26).alignment = Alignment(horizontal="center")
sheet.cell(2, 27).alignment = Alignment(horizontal="center")
sheet.cell(2, 28).alignment = Alignment(horizontal="center")
sheet.cell(2, 29).alignment = Alignment(horizontal="center")
sheet.cell(2, 30).alignment = Alignment(horizontal="center")
sheet.cell(2, 31).alignment = Alignment(horizontal="center")
sheet.cell(2, 32).alignment = Alignment(horizontal="center")
sheet.cell(2, 33).alignment = Alignment(horizontal="center")
sheet.cell(2, 34).alignment = Alignment(horizontal="center")
sheet.cell(2, 35).alignment = Alignment(horizontal="center")
sheet.cell(2, 36).alignment = Alignment(horizontal="center")
sheet.cell(2, 37).alignment = Alignment(horizontal="center")
sheet.cell(2, 38).alignment = Alignment(horizontal="center")
sheet.cell(2, 39).alignment = Alignment(horizontal="center")
sheet.cell(2, 40).alignment = Alignment(horizontal="center")
sheet.cell(2, 41).alignment = Alignment(horizontal="center")
sheet.cell(2, 42).alignment = Alignment(horizontal="center")
sheet.cell(2, 43).alignment = Alignment(horizontal="center")
sheet.cell(2, 44).alignment = Alignment(horizontal="center")
sheet.cell(2, 45).alignment = Alignment(horizontal="center")
sheet.cell(2, 46).alignment = Alignment(horizontal="center")
sheet.cell(2, 47).alignment = Alignment(horizontal="center")
sheet.cell(2, 48).alignment = Alignment(horizontal="center")
sheet.cell(2, 49).alignment = Alignment(horizontal="center")
sheet.cell(2, 50).alignment = Alignment(horizontal="center")
sheet.cell(2, 51).alignment = Alignment(horizontal="center")
sheet.cell(2, 52).alignment = Alignment(horizontal="center")
sheet.cell(2, 53).alignment = Alignment(horizontal="center")
sheet.cell(2, 54).alignment = Alignment(horizontal="center")
sheet.cell(2, 55).alignment = Alignment(horizontal="center")
sheet.cell(2, 3).border = thick_border
sheet.cell(2, 4).border = thick_border
sheet.cell(2, 5).border = thick_border
sheet.cell(2, 6).border = thick_border
sheet.cell(2, 7).border = thick_border
sheet.cell(2, 8).border = thick_border
sheet.cell(2, 9).border = thick_border
sheet.cell(2, 10).border = thick_border
sheet.cell(2, 11).border = thick_border
sheet.cell(2, 12).border = thick_border
sheet.cell(2, 13).border = thick_border
sheet.cell(2, 14).border = thick_border
sheet.cell(2, 15).border = thick_border
sheet.cell(2, 16).border = thick_border
sheet.cell(2, 17).border = thick_border
sheet.cell(2, 18).border = thick_border
sheet.cell(2, 19).border = thick_border
sheet.cell(2, 20).border = thick_border
sheet.cell(2, 21).border = thick_border
sheet.cell(2, 22).border = thick_border
sheet.cell(2, 23).border = thick_border
sheet.cell(2, 24).border = thick_border
sheet.cell(2, 25).border = thick_border
sheet.cell(2, 26).border = thick_border
sheet.cell(2, 27).border = thick_border
sheet.cell(2, 28).border = thick_border
sheet.cell(2, 29).border = thick_border
sheet.cell(2, 30).border = thick_border
sheet.cell(2, 31).border = thick_border
sheet.cell(2, 32).border = thick_border
sheet.cell(2, 33).border = thick_border
sheet.cell(2, 34).border = thick_border
sheet.cell(2, 35).border = thick_border
sheet.cell(2, 36).border = thick_border
sheet.cell(2, 37).border = thick_border
sheet.cell(2, 38).border = thick_border
sheet.cell(2, 39).border = thick_border
sheet.cell(2, 40).border = thick_border
sheet.cell(2, 41).border = thick_border
sheet.cell(2, 42).border = thick_border
sheet.cell(2, 43).border = thick_border
sheet.cell(2, 44).border = thick_border
sheet.cell(2, 45).border = thick_border
sheet.cell(2, 46).border = thick_border
sheet.cell(2, 47).border = thick_border
sheet.cell(2, 48).border = thick_border
sheet.cell(2, 49).border = thick_border
sheet.cell(2, 50).border = thick_border
sheet.cell(2, 51).border = thick_border
sheet.cell(2, 52).border = thick_border
sheet.cell(2, 53).border = thick_border
sheet.cell(2, 54).border = thick_border
sheet.cell(2, 55).border = thick_border

sheet.freeze_panes = 'B3'
row_counter = 3
frame_count = 0
print(range(int(len(weeklydata.index)/18)))
for i in range(int(len(weeklydata.index)/18)):
    if(frame_count < len(weeklydata.index)):
        first_frame = frame_count
        #step = int(i / 18)
        #print("%s row frame, %s sheet row, %s date row", (row_counter-(i*1)-3), row_counter, (((i)*20)+3))
        sheet.cell((((i)*20)+3), 1).value = "%s/%s/%s" % (weeklydata.iloc[frame_count]['Month'], weeklydata.iloc[frame_count]['Week Ending'], weeklydata.iloc[frame_count]['Year'])
        sheet.cell((((i)*20)+3), 1).border = thick_border
        for row in range(19):
            #print (row)
            sheet.cell(row_counter, 2).value = weeklydata.iloc[frame_count]['Dept']

            #Gross Sales
            sheet.cell(row_counter, 3).value = weeklydata.iloc[frame_count]['Revenue']
            sheet.cell(row_counter, 3).number_format = '"$"#,##0'
            sheet.cell(row_counter, 7).value = weeklydata.iloc[frame_count]['Ninety_Trailing_Revenue']
            sheet.cell(row_counter, 7).number_format = '"$"#,##0'
            sheet.cell(row_counter, 8).value = weeklydata.iloc[frame_count]['LY_REV']
            sheet.cell(row_counter, 8).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 9).value = weeklydata.iloc[frame_count]['REV_GRO']
            sheet.cell(row_counter, 9).number_format = '0.00%'
            if(weeklydata.iloc[frame_count]['REV_GRO']<0): sheet.cell(row_counter, 9).fill = redFill
            else: sheet.cell(row_counter, 9).fill = greenFill

            #COGS
            sheet.cell(row_counter, 10).value = weeklydata.iloc[frame_count]['Cost']
            sheet.cell(row_counter, 10).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 13).value = weeklydata.iloc[frame_count]['Ninety_Trailing_Cost']
            sheet.cell(row_counter, 13).number_format = '"$"#,##0'
            sheet.cell(row_counter, 14).value = weeklydata.iloc[frame_count]['LY_COST']
            sheet.cell(row_counter, 14).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 15).value = weeklydata.iloc[frame_count]['COST_GRO']
            sheet.cell(row_counter, 15).number_format = '0.00%'

            #Gross Margin
            sheet.cell(row_counter, 16).value = weeklydata.iloc[frame_count]['Margin']
            sheet.cell(row_counter, 16).number_format = '0.00%'
            sheet.cell(row_counter, 19).value = weeklydata.iloc[frame_count]['Ninety_Trailing_Margin']
            sheet.cell(row_counter, 19).number_format = '0.00%'
            sheet.cell(row_counter, 20).value = weeklydata.iloc[frame_count]['LY_MARGIN']
            sheet.cell(row_counter, 20).number_format = '0.00%'
            sheet.cell(row_counter, 21).value = weeklydata.iloc[frame_count]['MARGIN_GROWTH']
            sheet.cell(row_counter, 21).number_format = '0.00%'
            if(weeklydata.iloc[frame_count]['MARGIN_GROWTH']<0): sheet.cell(row_counter, 21).fill = redFill
            else: sheet.cell(row_counter, 21).fill = greenFill

            #Profit
            sheet.cell(row_counter, 22).value = weeklydata.iloc[frame_count]['Profit']
            sheet.cell(row_counter, 22).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 26).value = weeklydata.iloc[frame_count]['Ninety_Trailing_Profit']
            sheet.cell(row_counter, 26).number_format = '"$"#,##0'
            sheet.cell(row_counter, 27).value = weeklydata.iloc[frame_count]['LY_PROFIT']
            sheet.cell(row_counter, 27).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 28).value = weeklydata.iloc[frame_count]['PROFIT_GROWTH']
            sheet.cell(row_counter, 28).number_format = '0.00%'
            if(weeklydata.iloc[frame_count]['PROFIT_GROWTH']<0): sheet.cell(row_counter, 28).fill = redFill
            else: sheet.cell(row_counter, 28).fill = greenFill

            #Volume
            sheet.cell(row_counter, 29).value = weeklydata.iloc[frame_count]['Qty']
            sheet.cell(row_counter, 30).value = weeklydata.iloc[frame_count]['Ninety_Trailing_Volume']
            sheet.cell(row_counter, 31).value = weeklydata.iloc[frame_count]['LY_Volume']
            sheet.cell(row_counter, 32).value = weeklydata.iloc[frame_count]['VOL_GRO']

            #Variance
            sheet.cell(row_counter, 33).value = weeklydata.iloc[frame_count]['GrossProfitVariance']
            sheet.cell(row_counter, 33).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 34).value = weeklydata.iloc[frame_count]['DueToMargin']
            sheet.cell(row_counter, 34).number_format = '"$"#,##0_);[Red]("$"#,##0)'
            sheet.cell(row_counter, 35).value = weeklydata.iloc[frame_count]['VolumeVariance']
            sheet.cell(row_counter, 35).number_format = '"$"#,##0_);[Red]("$"#,##0)'


            sheet.cell(row_counter, 2).border = thick_border
            sheet.cell(row_counter, 3).border = thin_border
            sheet.cell(row_counter, 4).border = thin_border
            sheet.cell(row_counter, 5).border = thin_border
            sheet.cell(row_counter, 6).border = thin_border
            sheet.cell(row_counter, 7).border = thin_border
            sheet.cell(row_counter, 8).border = thin_border
            sheet.cell(row_counter, 9).border = thin_border
            sheet.cell(row_counter, 10).border = thin_border
            sheet.cell(row_counter, 11).border = thin_border
            sheet.cell(row_counter, 12).border = thin_border
            sheet.cell(row_counter, 13).border = thin_border
            sheet.cell(row_counter, 14).border = thin_border
            sheet.cell(row_counter, 15).border = thin_border
            sheet.cell(row_counter, 16).border = thin_border
            sheet.cell(row_counter, 17).border = thin_border
            sheet.cell(row_counter, 18).border = thin_border
            sheet.cell(row_counter, 19).border = thin_border
            sheet.cell(row_counter, 20).border = thin_border
            sheet.cell(row_counter, 21).border = thin_border
            sheet.cell(row_counter, 22).border = thin_border
            sheet.cell(row_counter, 23).border = thin_border
            sheet.cell(row_counter, 24).border = thin_border
            sheet.cell(row_counter, 25).border = thin_border
            sheet.cell(row_counter, 26).border = thin_border
            sheet.cell(row_counter, 27).border = thin_border
            sheet.cell(row_counter, 28).border = thin_border
            sheet.cell(row_counter, 29).border = thin_border
            sheet.cell(row_counter, 30).border = thin_border
            sheet.cell(row_counter, 31).border = thin_border
            sheet.cell(row_counter, 32).border = thin_border
            sheet.cell(row_counter, 33).border = thin_border
            sheet.cell(row_counter, 34).border = thin_border
            sheet.cell(row_counter, 35).border = thin_border
            sheet.cell(row_counter, 36).border = thin_border
            sheet.cell(row_counter, 37).border = thin_border
            sheet.cell(row_counter, 38).border = thin_border
            sheet.cell(row_counter, 39).border = thin_border
            sheet.cell(row_counter, 40).border = thin_border
            sheet.cell(row_counter, 41).border = thin_border
            sheet.cell(row_counter, 42).border = thin_border
            sheet.cell(row_counter, 43).border = thin_border
            sheet.cell(row_counter, 44).border = thin_border
            sheet.cell(row_counter, 45).border = thin_border
            sheet.cell(row_counter, 46).border = thin_border
            sheet.cell(row_counter, 47).border = thin_border
            sheet.cell(row_counter, 48).border = thin_border
            sheet.cell(row_counter, 49).border = thin_border
            sheet.cell(row_counter, 50).border = thin_border
            sheet.cell(row_counter, 51).border = thin_border
            sheet.cell(row_counter, 52).border = thin_border
            sheet.cell(row_counter, 53).border = thin_border
            sheet.cell(row_counter, 54).border = thin_border
            sheet.cell(row_counter, 55).border = thin_border

            frame_count+=1
            row_counter+=1

        last_frame = frame_count

        grand_rev = weeklydata[first_frame:last_frame]['Revenue'].sum()
        grand_cost = weeklydata[first_frame:last_frame]['Cost'].sum()
        grand_ly_rev = weeklydata[first_frame:last_frame]['LY_REV'].sum()
        grand_ly_cost = weeklydata[first_frame:last_frame]['LY_COST'].sum()
        grand_profit = grand_rev - grand_cost
        grand_ly_profit = grand_ly_rev - grand_ly_cost
        grand_margin = grand_profit / grand_rev
        grand_ly_margin = grand_ly_profit / grand_ly_rev
        grand_volume = weeklydata[first_frame:last_frame]['Qty'].sum()
        grand_ly_volume = weeklydata[first_frame:last_frame]['LY_Volume'].sum()
        grand_volume_growth = (grand_volume / grand_ly_volume) - 1
        grand_trailing_rev = weeklydata[first_frame:last_frame]['Ninety_Trailing_Revenue'].sum()
        grand_trailing_cost = weeklydata[first_frame:last_frame]['Ninety_Trailing_Cost'].sum()
        grand_trailing_volume = weeklydata[first_frame:last_frame]['Ninety_Trailing_Volume'].sum()
        grand_trailing_margin = weeklydata[first_frame:last_frame]['Ninety_Trailing_Margin'].sum()
        grand_trailing_profit = weeklydata[first_frame:last_frame]['Ninety_Trailing_Profit'].sum()

        #Gross Sales
        sheet.cell(row_counter, 2).value = 'GRAND TOTAL'
        sheet.cell(row_counter, 3).value = grand_rev
        sheet.cell(row_counter, 3).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 7).value = grand_trailing_rev
        sheet.cell(row_counter, 7).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 8).value = grand_ly_rev
        sheet.cell(row_counter, 8).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 9).value = (grand_rev - grand_ly_rev) / grand_ly_rev
        sheet.cell(row_counter, 9).number_format = '0.00%'
        if((grand_rev - grand_ly_rev) / grand_ly_rev<0): sheet.cell(row_counter, 9).fill = redFill
        else: sheet.cell(row_counter, 9).fill = greenFill

        #COGS
        sheet.cell(row_counter, 10).value = grand_cost
        sheet.cell(row_counter, 10).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 13).value = grand_trailing_cost
        sheet.cell(row_counter, 13).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 14).value = grand_ly_cost
        sheet.cell(row_counter, 14).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 15).value = (grand_cost - grand_ly_cost)/grand_ly_cost
        sheet.cell(row_counter, 15).number_format = '0.00%'

        #Gross Margin
        sheet.cell(row_counter, 16).value = grand_margin
        sheet.cell(row_counter, 16).number_format = '0.00%'
        sheet.cell(row_counter, 19).value = grand_trailing_margin
        sheet.cell(row_counter, 19).number_format = '0.00%'
        sheet.cell(row_counter, 20).value = grand_ly_margin
        sheet.cell(row_counter, 20).number_format = '0.00%'
        sheet.cell(row_counter, 21).value = (grand_margin - grand_ly_margin) / grand_ly_margin
        sheet.cell(row_counter, 21).number_format = '0.00%'
        if((grand_margin - grand_ly_margin) / grand_ly_margin<0): sheet.cell(row_counter, 21).fill = redFill
        else: sheet.cell(row_counter, 21).fill = greenFill

        #Profit
        sheet.cell(row_counter, 22).value = grand_profit
        sheet.cell(row_counter, 22).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 26).value = grand_trailing_profit
        sheet.cell(row_counter, 26).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 27).value = grand_ly_profit
        sheet.cell(row_counter, 27).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 28).value = (grand_profit - grand_ly_profit)/grand_ly_profit
        sheet.cell(row_counter, 28).number_format = '0.00%'
        if((grand_profit - grand_ly_profit)/grand_ly_profit<0): sheet.cell(row_counter, 28).fill = redFill
        else: sheet.cell(row_counter, 28).fill = greenFill

        #Volume
        sheet.cell(row_counter, 29).value = round(grand_volume, 1)
        sheet.cell(row_counter, 30).value = round(grand_trailing_volume, 1)
        sheet.cell(row_counter, 31).value = round(grand_ly_volume, 1)
        sheet.cell(row_counter, 32).value = round(grand_volume_growth, 1)

        #Variance
        sheet.cell(row_counter, 33).value = grand_profit - grand_ly_profit
        sheet.cell(row_counter, 33).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 34).value = (grand_rev * grand_margin) - (grand_rev * grand_ly_margin)
        sheet.cell(row_counter, 34).number_format = '"$"#,##0_);[Red]("$"#,##0)'
        sheet.cell(row_counter, 35).value = (grand_volume - grand_ly_volume) * (grand_cost / grand_volume)
        sheet.cell(row_counter, 35).number_format = '"$"#,##0_);[Red]("$"#,##0)'


        sheet.cell(row_counter, 2).border = thick_border
        sheet.cell(row_counter, 3).border = thick_border
        sheet.cell(row_counter, 4).border = thick_border
        sheet.cell(row_counter, 5).border = thick_border
        sheet.cell(row_counter, 6).border = thick_border
        sheet.cell(row_counter, 7).border = thick_border
        sheet.cell(row_counter, 8).border = thick_border
        sheet.cell(row_counter, 9).border = thick_border
        sheet.cell(row_counter, 10).border = thick_border
        sheet.cell(row_counter, 11).border = thick_border
        sheet.cell(row_counter, 12).border = thick_border
        sheet.cell(row_counter, 13).border = thick_border
        sheet.cell(row_counter, 14).border = thick_border
        sheet.cell(row_counter, 15).border = thick_border
        sheet.cell(row_counter, 16).border = thick_border
        sheet.cell(row_counter, 17).border = thick_border
        sheet.cell(row_counter, 18).border = thick_border
        sheet.cell(row_counter, 19).border = thick_border
        sheet.cell(row_counter, 20).border = thick_border
        sheet.cell(row_counter, 21).border = thick_border
        sheet.cell(row_counter, 22).border = thick_border
        sheet.cell(row_counter, 23).border = thick_border
        sheet.cell(row_counter, 24).border = thick_border
        sheet.cell(row_counter, 25).border = thick_border
        sheet.cell(row_counter, 26).border = thick_border
        sheet.cell(row_counter, 27).border = thick_border
        sheet.cell(row_counter, 28).border = thick_border
        sheet.cell(row_counter, 29).border = thick_border
        sheet.cell(row_counter, 30).border = thick_border
        sheet.cell(row_counter, 31).border = thick_border
        sheet.cell(row_counter, 32).border = thick_border
        sheet.cell(row_counter, 33).border = thick_border
        sheet.cell(row_counter, 34).border = thick_border
        sheet.cell(row_counter, 35).border = thick_border
        sheet.cell(row_counter, 36).border = thick_border
        sheet.cell(row_counter, 37).border = thick_border
        sheet.cell(row_counter, 38).border = thick_border
        sheet.cell(row_counter, 39).border = thick_border
        sheet.cell(row_counter, 40).border = thick_border
        sheet.cell(row_counter, 41).border = thick_border
        sheet.cell(row_counter, 42).border = thick_border
        sheet.cell(row_counter, 43).border = thick_border
        sheet.cell(row_counter, 44).border = thick_border
        sheet.cell(row_counter, 45).border = thick_border
        sheet.cell(row_counter, 46).border = thick_border
        sheet.cell(row_counter, 47).border = thick_border
        sheet.cell(row_counter, 48).border = thick_border
        sheet.cell(row_counter, 49).border = thick_border
        sheet.cell(row_counter, 50).border = thick_border
        sheet.cell(row_counter, 51).border = thick_border
        sheet.cell(row_counter, 52).border = thick_border
        sheet.cell(row_counter, 53).border = thick_border
        sheet.cell(row_counter, 54).border = thick_border
        sheet.cell(row_counter, 55).border = thick_border
        row_counter+=1

        for i in range(19):
            sheet.cell(row_counter-(i+2), 4).value = weeklydata.iloc[frame_count-(i+1)]['Revenue'] / grand_rev
            sheet.cell(row_counter-(i+2), 4).number_format = '0.00%'
            sheet.cell(row_counter-(i+2), 23).value = weeklydata.iloc[frame_count-(i+1)]['Profit'] / grand_profit
            sheet.cell(row_counter-(i+2), 23).number_format = '0.00%'

    

pd.save("financials-monthly.xlsx")

print(weeklydata)