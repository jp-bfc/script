#https://stackoverflow.com/questions/26951538/python-using-xlrd-to-obtain-column-heading-and-using-a-loop-to-create-variables

import xlrd
import csv
import openpyxl
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
import datetime
import pyodbc
import pandas as pd
from ftplib import FTP
from pathlib import Path
import configparser


class Item:
    def __init__(self, row = []):
        self.ITEM_UPC = row[0],
        self.Brand = row[1],
        self.Description = row[2],
        self.Size = row[3],
        self.Department = row[4],
        self.Sub_Dept = row[5],
        self.Local = row[6],
        self.Organic = row[7],
        self.POS_Desc = row[8],
        self.Bottle_Link = row[9],
        self.Link_Val = row[10],
        self.SellSize = row[11],
        self.UoM = row[12],
        self.BASICS = row[13],
        self.Weight_Qty = row[14],
        self.Promo_Code = row[15],
        self.VT_Tax = row[16],
        self.Meals_Tax = row[17],
        self.Local2 = row[18],
        self.Org_Local = row[19],
        self.FT = row[20],
        self.Coop = row[21],
        self.Tax = row[22],
        self.PLU = row[23],
        self.WIC = row[24],
        self.F180 = row[25],
        self.V = row[26],
        self.Six = row[27],
        self.M = row[28],
        self.F = row[29],
        self.C = row[30],
        self.O = row[31],
        self.W = row[32],
        self.B = row[33],
        self.N = row[34],
        self.G = row[35],
        self.A = row[36],
        self.E = row[37],
        self.Scaled = row[38],
        self.Vnd_ID = row[39],
        self.Vnd_Code = row[40],
        self.Case_Cost = row[41],
        self.Units_Per_Case = row[42],
        self.Unit_Cost = row[43],
        self.Price = row[44],
        self.Sale_Price = row[45],
        self.Sale_Start = row[46],
        self.Sale_End = row[47]

previousDate = (datetime.datetime.today() - datetime.timedelta(days=1)).date()
pdFormatted = previousDate.strftime('%m%d%Y')
#https://stackoverflow.com/questions/9319317/quick-and-easy-file-dialog-in-python
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

class Row:
    Cost = float
    Price = float
    Qty = float
    Revenue = float
    def __init__(self, UPC, DESC, Cost, Price, Qty, Revenue):
        self.UPC = UPC
        self.DESC = DESC
        self.Cost = Cost
        self.Price = Price
        self.Qty = Qty
        self.Revenue = Revenue
    
    def __lt__(self, other):
        return self.Revenue < other.Revenue

#https://stackoverflow.com/questions/2262333/is-there-a-built-in-or-more-pythonic-way-to-try-to-parse-a-string-to-an-integer
def intTryParse(value):
    try:
        return int(value), True
    except ValueError:
        return value, False

#TRY THIS FIRST
#select w.F01 as 'UPC', o.F29 as 'DESCRIPTION', c.F1140 as 'COST', w.F65 as 'PRICE', w.F64 as 'QTY'
#from (select * from STORESQL.dbo.RPT_ITM_W where F254 = '2022-07-03 00:00:00.000' and F1034 = 3) w
#inner join STORESQL.dbo.OBJ_TAB o on w.F01 = o.F01
#inner join STORESQL.dbo.COST_TAB c on o.F01 = c.f01
#inner join STORESQL.dbo.PRICE_TAB p on w.F01 = p.F01


with xlrd.open_workbook(file_path) as wb:
    sh = wb.sheet_by_index(0)

    headers = {
        "Code":0,
        "Descriptor":1,
        "Unit Sold":5,
        "Unit Price":6,
        "Revenue":7,
        "Unit Cost":8
    }

    arr = list()

    for rowind in range(sh.nrows)[6:]:
        if(intTryParse(sh.row(rowind)[0].value)[1]):
            arr.append(Row(
                str(sh.row(rowind)[headers.get("Code")].value).rstrip('0').rstrip('.') if '.' in str(sh.row(rowind)[headers.get("Code")].value) else str(sh.row(rowind)[headers.get("Code")].value),
                sh.row(rowind)[headers.get("Descriptor")].value,
                sh.row(rowind)[headers.get("Unit Cost")].value if len(str(sh.row(rowind)[headers.get("Unit Cost")].value)) > 0 else 0,
                sh.row(rowind)[headers.get("Unit Price")].value if len(str(sh.row(rowind)[headers.get("Unit Price")].value)) > 0 else 0,
                sh.row(rowind)[headers.get("Unit Sold")].value if len(str(sh.row(rowind)[headers.get("Unit Sold")].value)) > 0 else 0,
                sh.row(rowind)[headers.get("Revenue")].value if len(str(sh.row(rowind)[headers.get("Revenue")].value)) > 0 else 0,
            ))

    #for key,value in dataDict.items():
     #   print ("UPC: %(u)s DESC: %(d)s Cost: %(c)s Price: %(p)s Qty: %(q)s Rev: %(r)s" % {'u':value.UPC, 'd':value.DESC, 'c':value.Cost, 'p':value.Price, 'q':value.Qty, 'r':value.Revenue })
    fn = "bfc_%s.csv" % pdFormatted

    file = Path('//bfc-hv-01/SWAP/New Reports/%s' % fn)
    file.touch(exist_ok=True)
        
    with open(file, 'w', newline="\n") as f:
        c = csv.writer(f)
        c.writerow(("UPC", "DESCRIPTION", "COST", "PRICE", "QTY", "REVENUE"))
        for value in arr:
            c.writerow((value.UPC, value.DESC, value.Cost, value.Price, value.Qty, value.Revenue))

    ftp = FTP('ftp.spins.com')
    ftp.login(user='brattleboro_food_coop', passwd='lm8uX0vZd')
    filename = file.name
    #path = Path('//bfc-hv-01/SWAP/New Reports/Completed/%s' % filename)

    ftpResponseMessage = ftp.storbinary('STOR '+filename, open(file, 'rb'))
    print(ftpResponseMessage)
    ftp.quit()

    arr.sort(reverse = True)

    bookDate = "Top 100 - %s.xlsx" % pdFormatted
    bookFile = Path('//bfc-hv-01/SWAP/New Reports/%s' % bookDate)
    bookFile.touch(exist_ok=True)

    pb = openpyxl.Workbook()

    pb.create_sheet("Top 100 %s" % pdFormatted)
    sheet = pb['Top 100 %s' % pdFormatted]

    sheet.cell(row=1, column=1).value = "UPC"
    sheet.cell(row=1, column=2).value = "DESCRIPTION"
    sheet.cell(row=1, column=3).value = "COST"
    sheet.cell(row=1, column=4).value = "PRICE"
    sheet.cell(row=1, column=5).value = "QTY"
    sheet.cell(row=1, column=6).value = "REVENUE"
    sheet.cell(row=1, column=7).value = "DEPT"
    sheet.cell(row=1, column=8).value = "SUB DEPT"
    sheet.cell(row=1, column=9).value = "PROFIT"
    sheet.cell(row=1, column=10).value = "MARGIN"

    rowIndex = 2

    config = configparser.ConfigParser()
    config.read('script_configs.ini')
    username = config['DEFAULT']['user']
    password = config['DEFAULT']['password']
    server = config['DEFAULT']['server']
    port = config['DEFAULT']['port']
    database = config['DEFAULT']['database']
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';PORT='+port+';DATABASE='+database+';UID='+username+';PWD='+ password)
    cursor = cnxn.cursor()
    
    query = """SELECT rpc.F1024 as 'Department', sdp.F1022 as 'SubDept'
     FROM STORESQL.dbo.OBJ_TAB OBJ
     LEFT join STORESQL.dbo.RPC_TAB rpc on OBJ.F18 = rpc.F18
     LEFT join POS_TAB pos on OBJ.F01 = pos.F01
     LEFT join STORESQL.dbo.SDP_TAB sdp on pos.F04 = sdp.F04
     WHERE OBJ.F01 = ?"""
    #str(cursor.execute("SELECT  FROM STORESQL.dbo.OBJ_TAB OBJ  WHERE OBJ.F01 = ?",  upc).fetchone()[0]).format('c')

    for value in arr:
        upc = ('0' * (13 - len(value.UPC))) + str((value.UPC))
        profit = (value.Revenue - (float(value.Cost) * float(value.Qty)))
        data = cursor.execute(query, str(upc)).fetchone()
        sheet.cell(row=rowIndex, column=1).value = upc
        sheet.cell(row=rowIndex, column=2).value = value.DESC
        sheet.cell(row=rowIndex, column=3).value = value.Cost
        sheet.cell(row=rowIndex, column=3).number_format = '$#,##0.00' 
        sheet.cell(row=rowIndex, column=4).value = value.Price
        sheet.cell(row=rowIndex, column=4).number_format = '$#,##0.00' 
        sheet.cell(row=rowIndex, column=5).value = value.Qty
        sheet.cell(row=rowIndex, column=6).value = value.Revenue
        sheet.cell(row=rowIndex, column=6).number_format = '$#,##0.00' 
        sheet.cell(row=rowIndex, column=7).value = data[0]
        sheet.cell(row=rowIndex, column=8).value = data[1]
        sheet.cell(row=rowIndex, column=9).value = profit
        sheet.cell(row=rowIndex, column=9).number_format = '$#,##0.00'
        sheet.cell(row=rowIndex, column=10).value = (profit / value.Revenue) if value.Revenue > 0 else "0"
        sheet.cell(row=rowIndex, column=10).number_format = '%'
        rowIndex += 1

    pb.save(bookFile)
    pb.close()

    salesFrame = pd.read_excel(bookFile, "Top 100 %s" % pdFormatted)
    
    second = openpyxl.load_workbook(bookFile)

    second.create_sheet('Total Sales %s' % pdFormatted)
    sheet = second['Total Sales %s' % pdFormatted]
    salesFrame['TOTALCOST'] = salesFrame.COST * salesFrame.QTY

    summedFrame = salesFrame.groupby(['DEPT'])[['REVENUE', 'TOTALCOST', 'QTY']].sum().reset_index()

    sheet.cell(row=1, column=1).value = "DEPARTMENT"
    sheet.cell(row=1, column=2).value = "REVENUE"
    sheet.cell(row=1, column=3).value = "COST"
    sheet.cell(row=1, column=4).value = "VOLUME"
    sheet.cell(row=1, column=5).value = "PROFIT"
    sheet.cell(row=1, column=6).value = "GROSS MARGIN"
    sheet.cell(row=1, column=7).value = "% REVENUE"
    sheet.cell(row=1, column=8).value = "% PROFIT"

    rowIndex = 2
    totalRev = summedFrame['REVENUE'].sum()
    totalCost = summedFrame['TOTALCOST'].sum()
    totalProfit = totalRev - totalCost
    for index, value in summedFrame.iterrows():
        dept = value['DEPT']
        rev = value['REVENUE']
        cost = value['TOTALCOST']
        volume = value['QTY']
        profit = rev - cost
        sheet.cell(row=rowIndex, column=1).value = dept
        sheet.cell(row=rowIndex, column=2).value = rev
        sheet.cell(row=rowIndex, column=3).value = cost
        sheet.cell(row=rowIndex, column=4).value = volume
        sheet.cell(row=rowIndex, column=5).value = profit
        #Gross Margin
        sheet.cell(row=rowIndex, column=6).value = profit / rev if rev > 0 else "0"
        #Percent of Total Store Revenue for Dept
        sheet.cell(row=rowIndex, column=7).value = rev / totalRev if rev > 0 else "0"
        #Percent of Total Store Profit for Dept
        sheet.cell(row=rowIndex, column=8).value = profit / totalProfit if rev > 0 else "0"
        rowIndex += 1

    sheet.cell(row=rowIndex, column=1).value = "GRAND TOTAL"
    sheet.cell(row=rowIndex, column=2).value = totalRev
    sheet.cell(row=rowIndex, column=2).value = totalCost
    sheet.cell(row=rowIndex, column=5).value = totalProfit
    second.save(bookFile)