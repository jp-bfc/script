import pandas as pd
import openpyxl as op
import tkinter as tk
from tkinter import filedialog
from pathlib import Path
import datetime
import os

root = tk.Tk()
root.withdraw()
ls = []
files = filedialog.askopenfilenames(parent=root, title='Choose labor files')
for file_path in files:
    # f = file_path[-15:-5]
    # print(f)
    #Quarter
    week = os.path.split(file_path)[1].split()[0]
    #Week
    #week = os.path.split(file_path)[1].split()[-1].split(".")[0]
    print(week)
    wb = op.load_workbook(filename=file_path)
    
    ws = wb.active
    for i in range(1, ws.max_row):
        if ws.cell(i, 1).value is not None:
            cell_val = ws.cell(i, 1).value
            if "Department:" in cell_val.split() and "Totals" not in cell_val.split():
                dept = cell_val.split()[1]
                lookup_row = i
                totalAmount = 0
                totalHours = 0
                while ws.cell(lookup_row, 18).value != "ER Liab":
                    if ws.cell(lookup_row, 3).value in ["REG", "SALRY", "OT"]:
                        totalAmount += float(ws.cell(lookup_row, 6).value)
                        totalHours += float(ws.cell(lookup_row, 5).value)
                    lookup_row+=1
                ls.append([dept, totalAmount, totalHours, ws.cell(lookup_row, 5).value, ws.cell(lookup_row, 19).value, week])
            elif "Report" in cell_val.split():
                dept = "Total Store"
                lookup_row = i
                totalAmount = 0
                totalHours = 0
                while ws.cell(lookup_row, 18).value != "ER Liab":
                    if ws.cell(lookup_row, 3).value in ["REG", "SALRY", "OT"]:
                        totalAmount += float(ws.cell(lookup_row, 6).value)
                        totalHours += float(ws.cell(lookup_row, 5).value)
                    lookup_row+=1
                ls.append([dept, totalAmount, totalHours, ws.cell(lookup_row, 5).value, ws.cell(lookup_row, 19).value, week])
frame = pd.DataFrame(ls, columns=['Dept', "Worked Wages", "Worked Hours", 'Hours Total', 'Labor Total', 'Week'])
frame = frame.sort_values(['Week'], ascending=False)
frame.to_csv("Labor Quarterly Total %s.csv" % datetime.datetime.today().date(), index=False)
